from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database.models import Report
from locales import t
from utils.formatters import TASHKENT_TZ, UTC, format_report_type

# (tarjima kaliti, ustun kengligi)
COLUMNS: list[tuple[str, int]] = [
    ("col_no", 5),
    ("col_date", 12),
    ("col_time", 8),
    ("col_employee", 26),
    ("col_position", 22),
    ("col_type", 20),
    ("col_district", 22),
    ("col_customer", 30),
    ("col_partners", 28),
    ("col_plots", 12),
    ("col_status", 16),
    ("col_photos", 10),
    ("col_location", 24),
]

HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")

# Excel/openpyxl bu belgilar bilan boshlangan matnni formula deb hisoblaydi
# (masalan, "=cmd|'/c calc'!A1"). Hodim kiritgan matn (customer_name va h.k.)
# xujjatga to'g'ridan-to'g'ri tushgani uchun formula in'ektsiyasidan himoya kerak.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value):
    if isinstance(value, str) and value.startswith(_FORMULA_TRIGGERS):
        return "'" + value
    return value


def _local(dt):
    return (dt if dt.tzinfo else dt.replace(tzinfo=UTC)).astimezone(TASHKENT_TZ)


def _status_key(report: Report) -> str:
    if report.is_confirmed:
        return "status_confirmed"
    if report.is_rejected:
        return "status_rejected"
    return "status_pending"


def file_name(date_from: date, date_to: date) -> str:
    if date_from == date_to:
        return f"davomat-{date_from:%Y-%m-%d}.xlsx"
    return f"davomat-{date_from:%Y-%m-%d}_{date_to:%Y-%m-%d}.xlsx"


def build_workbook(lang: str, reports: list[Report]) -> bytes:
    """Hisobotlardan .xlsx fayl yasaydi va baytlarini qaytaradi.

    Fayl diskka yozilmaydi — Telegram ga to'g'ridan-to'g'ri baytlar
    yuboriladi, shuning uchun vaqtinchalik fayllarni tozalash muammosi
    ham yo'q.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = t(lang, "export_sheet_title")

    sheet.append([t(lang, key) for key, _ in COLUMNS])

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    for index, report in enumerate(reports, start=1):
        created = _local(report.created_at)
        sheet.append(
            [
                index,
                created.strftime("%d.%m.%Y"),
                created.strftime("%H:%M"),
                _safe_cell(report.user.full_name),
                _safe_cell(report.user.position),
                # Dinamik kalit yasash o'rniga mavjud formatter — tarjima
                # kalitlarini tekshiradigan test dinamik nomni ko'ra olmaydi.
                format_report_type(lang, report.report_type),
                report.district.name_uz if lang == "uz" else report.district.name_ru,
                _safe_cell(report.customer_name),
                _safe_cell(", ".join(rp.partner.full_name for rp in report.partners)),
                report.plots_count,
                t(lang, _status_key(report)),
                len(report.photos),
                f"{report.location_lat:.6f}, {report.location_lon:.6f}",
            ]
        )

    for column, (_, width) in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    # Sarlavha doim ko'rinib tursin va ustunlar bo'yicha saralash bo'lsin
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{sheet.max_row}"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
