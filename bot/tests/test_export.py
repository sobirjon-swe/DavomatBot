from datetime import date, datetime, timedelta, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook

from database import crud
from database.models import ReportType
from handlers.admin.reports import format_range, parse_range
from utils.export import build_workbook, file_name


async def _report(session, author, district_id, *, customer="Buyurtmachi", created=None):
    report = await crud.create_report(
        session,
        user_id=author.id,
        report_type=ReportType.laboratory,
        district_id=district_id,
        customer_name=customer,
        location_lat=41.311081,
        location_lon=69.240562,
        plots_count=4,
    )
    await crud.add_report_photos(session, report.id, ["f1", "f2", "f3"])
    if created is not None:
        report.created_at = created
    await session.commit()
    return await crud.get_report_by_id(session, report.id)


def rows(payload: bytes):
    sheet = load_workbook(BytesIO(payload)).active
    return list(sheet.iter_rows(values_only=True))


# ─── Oraliqni o'qish ────────────────────────────────────────────────────────

TODAY = date(2026, 6, 15)


def test_oralik_oqiladi():
    assert parse_range("01.06.2026 - 30.06.2026", TODAY) == (
        date(2026, 6, 1),
        date(2026, 6, 30),
    )


def test_bitta_sana_bir_kunlik_oralik_beradi():
    assert parse_range("05.06.2026", TODAY) == (date(2026, 6, 5), date(2026, 6, 5))


def test_teskari_yozilgan_sanalar_almashtiriladi():
    """Foydalanuvchini xato bilan qaytarishdan ko'ra tushunib olgan ma'qul."""
    assert parse_range("30.06.2026 - 01.06.2026", TODAY) == (
        date(2026, 6, 1),
        date(2026, 6, 30),
    )


def test_uzun_tire_ham_qabul_qilinadi():
    assert parse_range("01.06.2026 — 05.06.2026", TODAY) is not None


@pytest.mark.parametrize(
    "raw",
    ["", "kecha", "2026-06-01", "01.06.2026 - 05.06.2026 - 10.06.2026", "32.13.2026"],
)
def test_notogri_format_rad_etiladi(raw):
    assert parse_range(raw, TODAY) is None


def test_oralik_korinishi():
    assert format_range(date(2026, 6, 5), date(2026, 6, 5)) == "05.06.2026"
    assert format_range(date(2026, 6, 1), date(2026, 6, 30)) == "01.06.2026 - 30.06.2026"


# ─── Fayl nomi ──────────────────────────────────────────────────────────────

def test_fayl_nomi_bir_kun():
    assert file_name(date(2026, 6, 5), date(2026, 6, 5)) == "davomat-2026-06-05.xlsx"


def test_fayl_nomi_oralik():
    assert (
        file_name(date(2026, 6, 1), date(2026, 6, 30))
        == "davomat-2026-06-01_2026-06-30.xlsx"
    )


# ─── Kitob mazmuni ──────────────────────────────────────────────────────────

async def test_sarlavha_va_satrlar(session, employee, districts):
    await _report(session, employee, districts[0].id, customer="Birinchi")
    await _report(session, employee, districts[0].id, customer="Ikkinchi")
    reports = await crud.list_reports(session)

    table = rows(build_workbook("uz", reports))

    assert table[0][:4] == ("№", "Sana", "Vaqt", "Hodim")
    assert len(table) == 3  # sarlavha + 2 ta hisobot
    assert {table[1][7], table[2][7]} == {"Birinchi", "Ikkinchi"}


async def test_qatorlar_toliq_toladi(session, employee, districts):
    await _report(session, employee, districts[0].id)
    reports = await crud.list_reports(session)

    row = rows(build_workbook("uz", reports))[1]

    assert row[0] == 1
    assert row[3] == "Ali Valiyev"
    assert row[4] == "Muhandis"
    assert row[5] == "Laboratoriya"
    assert row[7] == "Buyurtmachi"
    assert row[9] == 4  # uchastkalar
    assert row[10] == "Kutilmoqda"
    assert row[11] == 3  # rasmlar
    assert row[12].startswith("41.311081")


async def test_vaqt_toshkent_mintaqasida(session, employee, districts):
    """Regressiya: bazada UTC turadi, faylda mahalliy vaqt bo'lishi kerak."""
    await _report(
        session,
        employee,
        districts[0].id,
        created=datetime(2026, 6, 5, 6, 30, tzinfo=timezone.utc),
    )
    reports = await crud.list_reports(session)

    row = rows(build_workbook("uz", reports))[1]

    assert row[1] == "05.06.2026"
    assert row[2] == "11:30"


async def test_tasdiqlangan_holat_korinadi(session, employee, admin, districts):
    report = await _report(session, employee, districts[0].id)
    await crud.confirm_report(session, report.id, admin.id)
    reports = await crud.list_reports(session)

    assert rows(build_workbook("uz", reports))[1][10] == "Tasdiqlangan"


async def test_sheriklar_royxati(session, employee, admin, districts):
    report = await _report(session, employee, districts[0].id)
    await crud.add_report_partners(session, report.id, [admin.id])
    await session.commit()

    # expire_all shart: obyekt sessiya keshida sheriklarsiz yotibdi va
    # qayta so'rov ham o'sha eski nusxani qaytarardi.
    session.expire_all()
    reports = await crud.list_reports(session)

    assert rows(build_workbook("uz", reports))[1][8] == "Salim Adminov"


async def test_ruscha_sarlavhalar(session, employee, districts):
    await _report(session, employee, districts[0].id)
    reports = await crud.list_reports(session)

    table = rows(build_workbook("ru", reports))

    assert table[0][3] == "Сотрудник"
    assert table[1][10] == "Ожидает"


def test_bosh_royxatda_faqat_sarlavha():
    assert len(rows(build_workbook("uz", []))) == 1


# ─── Oraliq bo'yicha tanlash ────────────────────────────────────────────────

async def test_oralik_chekkalari_ham_kiradi(session, employee, districts):
    bugun = crud.local_today()
    await _report(session, employee, districts[0].id, customer="Bugun")

    kechagi = await _report(session, employee, districts[0].id, customer="Kecha")
    kechagi.created_at = kechagi.created_at - timedelta(days=1)
    await session.commit()

    faqat_bugun = await crud.list_reports(session, date_from=bugun, date_to=bugun)
    ikki_kun = await crud.list_reports(
        session, date_from=bugun - timedelta(days=1), date_to=bugun
    )

    assert [r.customer_name for r in faqat_bugun] == ["Bugun"]
    assert len(ikki_kun) == 2
